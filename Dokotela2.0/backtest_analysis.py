import warnings; warnings.filterwarnings("ignore")
import yfinance as yf, pandas as pd, pandas_ta as ta, numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

PAIRS={"EURUSD":"EURUSD=X","AUDUSD":"AUDUSD=X","USDJPY":"USDJPY=X","GBPUSD":"GBPUSD=X","AUDCAD":"AUDCAD=X"}
JPY={"USDJPY","EURJPY","GBPJPY","AUDJPY"}
MLEN=75;EF=9;ES=21;SLM=1.0;ZB=1.0;MT=1;CD=5;RR=3.0;H1N=10

def fill(c): return PatternFill("solid",fgColor=c)
def fnt(c,b=False,s=11): return Font(color=c,bold=b,size=s)
def bdr():
    x=Side(style="thin",color="CCCCCC"); return Border(left=x,right=x,top=x,bottom=x)

def fetch(ticker,tf="15m",days=58):
    end=datetime.utcnow(); start=end-timedelta(days=days)
    df=yf.download(ticker,start=start,end=end,interval=tf,progress=False,auto_adjust=True)
    if df.empty: return df
    df.columns=[c[0].lower() if isinstance(c,tuple) else c.lower() for c in df.columns]
    df=df[[c for c in ["open","high","low","close"] if c in df.columns]].dropna()
    df.index=pd.to_datetime(df.index,utc=True)
    return df

def ind(df):
    d=df.copy()
    d["ef"]=ta.ema(d["close"],length=EF); d["es"]=ta.ema(d["close"],length=ES)
    d["atr"]=ta.atr(d["high"],d["low"],d["close"],length=14)
    d["sh"]=d["high"].rolling(MLEN).max(); d["sl"]=d["low"].rolling(MLEN).min()
    return d.dropna()

def h1_structure(df):
    n=H1N; h=df["high"].values; l=df["low"].values; idx=df.index
    bull_c=0; bear_c=0; bull_on=False; bear_on=False; ready={}
    for i in range(n,len(df)-n):
        t=idx[i]
        if h[i]==max(h[i-n:i+n+1]): bear_c+=1; bull_c=0
        elif l[i]==min(l[i-n:i+n+1]): bull_c+=1; bear_c=0
        if bull_c>=2: bull_on=True
        if bear_c>=2: bear_on=True
        d=set()
        if bull_on: d.add("BUY")
        if bear_on: d.add("SELL")
        ready[t]=d
    return ready

def h1_ok(etime,direction,structure):
    for t,dirs in structure.items():
        if t<etime and direction in dirs: return True
    return False

def run_signals(df,pair,h1s=None):
    df=ind(df).reset_index()
    # normalise the datetime column name to "time"
    time_col=[c for c in df.columns if "time" in str(c).lower() or "date" in str(c).lower()]
    if time_col and time_col[0]!="time": df=df.rename(columns={time_col[0]:"time"})
    elif "index" in df.columns: df=df.rename(columns={"index":"time"})
    mult=100 if pair in JPY else 10000
    trades=[]; sh1=sl1=np.nan; sht=slt=0; insh=insl=False
    ba=bea=False; intrade=False; tdir=ten=tsl=ttp=ti=None; cdl=0
    for i in range(1,len(df)):
        r=df.iloc[i]; p=df.iloc[i-1]; atr=r["atr"]
        if r["sh"]!=p["sh"]:
            if np.isnan(sh1) or abs(r["sh"]-sh1)>atr*0.5: sh1=r["sh"];sht=0;insh=False
        if r["sl"]!=p["sl"]:
            if np.isnan(sl1) or abs(r["sl"]-sl1)>atr*0.5: sl1=r["sl"];slt=0;insl=False
        if np.isnan(sh1): sh1=r["sh"]
        if np.isnan(sl1): sl1=r["sl"]
        nb=sl1-atr*ZB<=r["low"]<=sl1+atr*ZB
        nbe=sh1-atr*ZB<=r["high"]<=sh1+atr*ZB
        if nb and not insl: slt+=1;insl=True
        if not nb: insl=False
        if nbe and not insh: sht+=1;insh=True
        if not nbe: insh=False
        vb=slt>=MT; vbe=sht>=MT
        if intrade:
            hsl=(tdir=="BUY" and r["low"]<=tsl) or (tdir=="SELL" and r["high"]>=tsl)
            htp=(tdir=="BUY" and r["high"]>=ttp) or (tdir=="SELL" and r["low"]<=ttp)
            if hsl or htp:
                out="TP hit" if htp else "SL hit"; ep=ttp if htp else tsl
                pnl=((ep-ten) if tdir=="BUY" else (ten-ep))*mult
                trades[-1].update({"exit_time":str(r["time"])[:16],"exit_price":round(ep,5),
                    "outcome":out,"pnl_pips":round(pnl,1),"bars_held":i-ti})
                intrade=False
                if out=="SL hit": cdl=CD
        if cdl>0: cdl-=1
        if vb and nb and cdl==0: ba=True
        if vbe and nbe and cdl==0: bea=True
        if not nb or not vb: ba=False
        if not nbe or not vbe: bea=False
        bc=p["ef"]<p["es"] and r["ef"]>r["es"]; bec=p["ef"]>p["es"] and r["ef"]<r["es"]
        bb=r["ef"]>r["es"]; beb=r["ef"]<r["es"]
        buy=ba and bc and bb and not intrade
        sel=bea and bec and beb and not intrade
        if h1s is not None:
            et=r["time"]
            if buy: buy=h1_ok(et,"BUY",h1s)
            if sel: sel=h1_ok(et,"SELL",h1s)
        if buy:
            en=r["close"];s=sl1-atr*SLM;t2=sh1 if sh1>en else en+(en-s)*RR
            trades.append({"entry_time":str(r["time"])[:16],"direction":"BUY",
                "entry_price":round(en,5),"sl":round(s,5),"tp":round(t2,5),
                "swing_low":round(sl1,5),"swing_high":round(sh1,5),
                "exit_time":None,"exit_price":None,"outcome":"Open","pnl_pips":None,"bars_held":None})
            intrade=True;tdir="BUY";ten=en;tsl=s;ttp=t2;ti=i;ba=False
        elif sel:
            en=r["close"];s=sh1+atr*SLM;t2=sl1 if sl1<en else en-(s-en)*RR
            trades.append({"entry_time":str(r["time"])[:16],"direction":"SELL",
                "entry_price":round(en,5),"sl":round(s,5),"tp":round(t2,5),
                "swing_low":round(sl1,5),"swing_high":round(sh1,5),
                "exit_time":None,"exit_price":None,"outcome":"Open","pnl_pips":None,"bars_held":None})
            intrade=True;tdir="SELL";ten=en;tsl=s;ttp=t2;ti=i;bea=False
    return trades[:500]

def calc_stats(trades,pair):
    cl=[t for t in trades if t["outcome"]!="Open"]
    w=[t for t in cl if t["outcome"]=="TP hit"]; l=[t for t in cl if t["outcome"]=="SL hit"]
    wp=sum(t["pnl_pips"] for t in w if t["pnl_pips"]); lp=sum(t["pnl_pips"] for t in l if t["pnl_pips"])
    wr=round(len(w)/len(cl)*100,1) if cl else 0
    pf=round(wp/abs(lp),2) if lp else ("inf" if wp>0 else 0)
    return {"pair":pair,"total":len(trades),"closed":len(cl),"wins":len(w),"losses":len(l),
            "win_rate":wr,"pf":pf,"total_pips":round(wp+lp,1),
            "avg_win":round(wp/len(w),1) if w else 0,"avg_loss":round(lp/len(l),1) if l else 0}

COLS=[("Entry Time",18),("Direction",10),("Entry Price",13),("SL",11),("TP",11),
      ("Swing Low",12),("Swing High",12),("Exit Time",18),("Exit Price",13),
      ("Outcome",12),("P&L (pips)",13),("Bars Held",11),("R:R",8)]

def pair_sheet(wb,pair,trades,label,hdr_bg):
    ws=wb.create_sheet(title=f"{pair}-{label}"[:31])
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    b=ws["A1"]; b.value=f"{pair} — {label}"; b.fill=fill(hdr_bg)
    b.font=fnt("222222",b=True,s=12); b.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=24
    for ci,(name,w) in enumerate(COLS,1):
        c=ws.cell(row=2,column=ci,value=name); c.fill=fill("1E3A5F")
        c.font=fnt("FFFFFF",b=True); c.alignment=Alignment(horizontal="center"); c.border=bdr()
        ws.column_dimensions[get_column_letter(ci)].width=w
    for ri,t in enumerate(trades,3):
        risk=abs(t["entry_price"]-t["sl"]) if t["sl"] else 0
        rwd=abs(t["tp"]-t["entry_price"]) if t["tp"] else 0
        rr=round(rwd/risk,2) if risk>0 else ""
        row=[t["entry_time"],t["direction"],t["entry_price"],t["sl"],t["tp"],
             t["swing_low"],t["swing_high"],t["exit_time"] or "",t["exit_price"],
             t["outcome"],t["pnl_pips"],t["bars_held"],rr]
        for ci,v in enumerate(row,1):
            c=ws.cell(row=ri,column=ci,value=v); c.border=bdr(); c.alignment=Alignment(horizontal="center")
            if t["outcome"]=="TP hit": c.fill=fill("C6EFCE"); c.font=fnt("276221")
            elif t["outcome"]=="SL hit": c.fill=fill("FFC7CE"); c.font=fnt("9C0006")
            elif t["outcome"]=="Open": c.fill=fill("FFEB9C"); c.font=fnt("9C6500")
            elif ri%2==0: c.fill=fill("F2F7FF")
    s=calc_stats(trades,pair); sc=len(COLS)+2
    srows=[("STATISTICS","",True),("Total Trades",s["total"],False),("Closed",s["closed"],False),
           ("Wins",s["wins"],False),("Losses",s["losses"],False),("Win Rate",f"{s['win_rate']}%",False),
           ("Profit Factor",s["pf"],False),("Total P&L (pips)",s["total_pips"],False),
           ("Avg Win (pips)",s["avg_win"],False),("Avg Loss (pips)",s["avg_loss"],False)]
    for ri2,(lbl,val,hdr) in enumerate(srows,2):
        for ci2,v in enumerate([lbl,val],sc):
            c=ws.cell(row=ri2,column=ci2,value=v); c.border=bdr(); c.alignment=Alignment(horizontal="center")
            if hdr: c.fill=fill("2E4057"); c.font=fnt("FFFFFF",b=True)
            elif ri2%2==0: c.fill=fill("F2F7FF")
    ws.column_dimensions[get_column_letter(sc)].width=20
    ws.column_dimensions[get_column_letter(sc+1)].width=14
    ws.freeze_panes="A3"
    return s

def pf_col(pf):
    try:
        v=float(str(pf).replace("inf","999"))
        if v>=2.0: return "C6EFCE","276221"
        if v>=1.5: return "E2EFDA","375623"
        if v>=1.25: return "FFEB9C","9C6500"
        if v>=1.0: return "FCE4D6","833C00"
        return "FFC7CE","9C0006"
    except: return "FFFFFF","000000"

def summary_sheet(wb,d1,d2):
    ws=wb.create_sheet(title="SUMMARY",index=0)
    ws.merge_cells("A1:M1"); t=ws["A1"]; t.value="DOKOTELA 2.0 — DUAL DATASET ANALYSIS"
    t.fill=fill("2E4057"); t.font=Font(color="FFFFFF",bold=True,size=14)
    t.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=32
    ws.merge_cells("A2:M2"); s=ws["A2"]
    s.value=(f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC  |  "
             "D1=Base Strategy  |  D2=Base + H1 Structure Filter (2 pivots confirmed, entry on 3rd touch)  |  "
             f"Cooldown={CD} bars after SL  |  M15 timeframe")
    s.font=Font(color="555555",italic=True,size=9); s.alignment=Alignment(horizontal="center")
    ws.merge_cells("A3:M3"); g=ws["A3"]
    g.value=("  PF Guide:  <1.0 Losing  |  1.0-1.25 Marginal  |  1.25-1.5 Acceptable  |  "
             "1.5-2.0 Good  |  2.0-3.0 Very Good  |  >3.0 Excellent")
    g.fill=fill("FFF3CD"); g.font=Font(color="856404",size=10,italic=True)
    g.alignment=Alignment(horizontal="center"); ws.row_dimensions[3].height=16
    ws.merge_cells("B5:G5"); ws.merge_cells("H5:M5")
    for cell,lbl,bg in [("B5","DATASET 1 — Base Strategy","2980B9"),
                         ("H5","DATASET 2 — H1 Structure Filter (3rd Touch)","D35400")]:
        c=ws[cell]; c.value=lbl; c.fill=fill(bg); c.font=Font(color="FFFFFF",bold=True,size=11)
        c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[5].height=22
    hdrs=["Pair","Trades","Wins","Losses","Win%","PF","P&L(pips)"]; wids=[12,9,8,9,9,9,14]
    for offset in [1,7]:
        for ci,(h,w) in enumerate(zip(hdrs,wids),offset):
            c=ws.cell(row=6,column=ci,value=h)
            c.fill=fill("2980B9" if offset==1 else "D35400"); c.font=fnt("FFFFFF",b=True)
            c.alignment=Alignment(horizontal="center"); c.border=bdr()
            ws.column_dimensions[get_column_letter(ci)].width=w
    pairs=sorted(set([s["pair"] for s in d1+d2]))
    for ri,pair in enumerate(pairs,7):
        s1=next((s for s in d1 if s["pair"]==pair),None)
        s2=next((s for s in d2 if s["pair"]==pair),None)
        alt=ri%2==0
        for s,offset in [(s1,1),(s2,7)]:
            if not s:
                for ci in range(offset,offset+7):
                    c=ws.cell(row=ri,column=ci,value="—"); c.border=bdr(); c.alignment=Alignment(horizontal="center")
                continue
            pfbg,pffg=pf_col(s["pf"])
            row=[s["pair"],s["closed"],s["wins"],s["losses"],f"{s['win_rate']}%",s["pf"],s["total_pips"]]
            for ci,v in enumerate(row,offset):
                c=ws.cell(row=ri,column=ci,value=v); c.border=bdr(); c.alignment=Alignment(horizontal="center")
                col_pos=ci-offset
                if col_pos==5: c.fill=fill(pfbg); c.font=fnt(pffg,b=True)
                elif col_pos==6 and isinstance(v,(int,float)):
                    c.fill=fill("C6EFCE" if v>=0 else "FFC7CE"); c.font=fnt("276221" if v>=0 else "9C0006",b=True)
                elif alt: c.fill=fill("F2F7FF")
    tr=len(pairs)+7
    for s_list,offset,bg in [(d1,1,"2980B9"),(d2,7,"D35400")]:
        if not s_list: continue
        tc=sum(s["closed"] for s in s_list); tw=sum(s["wins"] for s in s_list)
        tl=sum(s["losses"] for s in s_list); tp=round(sum(s["total_pips"] for s in s_list),1)
        for ci,v in enumerate(["TOTAL",tc,tw,tl,f"{round(tw/max(tc,1)*100,1)}%","",tp],offset):
            c=ws.cell(row=tr,column=ci,value=v); c.fill=fill(bg)
            c.font=fnt("FFFFFF",b=True); c.alignment=Alignment(horizontal="center"); c.border=bdr()
    ws.merge_cells(f"A{tr+2}:M{tr+2}"); n=ws[f"A{tr+2}"]
    n.value=("  Green=TP hit  Red=SL hit  |  Compare D1 vs D2: higher PF with fewer trades in D2 = H1 filter improving quality.")
    n.fill=fill("EBF5FB"); n.font=Font(color="1A5276",italic=True,size=10)
    n.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[tr+2].height=18
    ws.freeze_panes="A7"

def main():
    wb=Workbook(); wb.remove(wb.active); d1=[]; d2=[]
    print("Running dual analysis...")
    for pair,ticker in PAIRS.items():
        print(f"  {pair}...",end=" ",flush=True)
        m15=fetch(ticker,"15m",58); h1=fetch(ticker,"60m",58)
        if m15.empty or len(m15)<MLEN+50: print("no data"); continue
        t1=run_signals(m15,pair); s1=pair_sheet(wb,pair,t1,"D1-Base","D6EAF8"); d1.append(s1)
        h1s={}
        if not h1.empty and len(h1)>=H1N*2+5: h1s=h1_structure(ind(h1))
        t2=run_signals(m15,pair,h1s=h1s); s2=pair_sheet(wb,pair,t2,"D2-H1Filter","FDEBD0"); d2.append(s2)
        print(f"D1:{s1['closed']} trades PF={s1['pf']} {s1['total_pips']:+.1f}pips  |  D2:{s2['closed']} trades PF={s2['pf']} {s2['total_pips']:+.1f}pips")
    if not d1: print("No data."); return
    summary_sheet(wb,d1,d2)
    fname=f"Dokotela2_DualAnalysis_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname); print(f"\nSaved: {fname}")
    print(f"\n{'Pair':<10}{'D1 Trades':>10}{'D1 WR':>8}{'D1 PF':>7}{'D1 P&L':>10}  |  {'D2 Trades':>10}{'D2 WR':>8}{'D2 PF':>7}{'D2 P&L':>10}")
    print("-"*85)
    for s1,s2 in zip(d1,d2):
        print(f"{s1['pair']:<10}{s1['closed']:>10}{s1['win_rate']:>7.1f}%{str(s1['pf']):>7}{s1['total_pips']:>+10.1f}  |  {s2['closed']:>10}{s2['win_rate']:>7.1f}%{str(s2['pf']):>7}{s2['total_pips']:>+10.1f}")

if __name__=="__main__":
    main()
